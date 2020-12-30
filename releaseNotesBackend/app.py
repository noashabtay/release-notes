from flask import Flask, abort, jsonify, request
from flask_cors import CORS
import pandas as pd
import openpyxl


# configuration
DEBUG = True
# instantiate the app
app = Flask(__name__)
app.config.from_object(__name__)


# enable CORS
CORS(app, resources={r'/*': {'origins': '*'}})


# sanity check route
@app.route('/ping', methods=['GET'])
def ping_pong():
    return jsonify('pong!')


@app.route('/releases', methods=['GET', 'POST'])
def insert_release():
    response_object = {}
    if request.method == 'POST':
        post_data = request.get_json()
        release = get_release_request(post_data)
        if insert_release_to_db(release):
            response_object['status'] = 'success'
            response_object['message'] = 'Release added!'
        else:
            abort(500)
    else:
        response_object['status'] = 'success'
        response_object['releases'] = init_versions()
    return jsonify(response_object)


@app.route('/releases/<release_id>', methods=['PUT'])
def update_single_releases(release_id):
    release = get_release_request(request.get_json())
    delete_release(release_id)
    if insert_release_to_db(release):
        response_object = {'status': 'success',
                           'message': 'Release updated!'}
    else:
        abort(500)
    return jsonify(response_object)


@app.route('/releases/<release_id>', methods=['DELETE'])
def delete_single_release(release_id):
    response_object = {'status': 'success'}
    delete_release(release_id)
    response_object['message'] = 'Release removed!'
    return jsonify(response_object)


@app.errorhandler(500)
def page_not_found(e):
    response_object = {'status': 'failed'}
    return response_object, 500


def find_last_row(release_id):
    """
    This function search for release id in excel file .
    return -1 if release_id exist
    return num of final row if doesnt exist
    :param release_id:
    :return:
    """
    wb = openpyxl.load_workbook('releasesData.xlsx')
    sheet = wb['Releases']
    i = 1
    while i <= sheet.max_row:
        if sheet.cell(row=i, column=1).value == release_id:
            wb.save('releasesData.xlsx')
            return -1
        elif sheet.cell(row=i, column=1).value is None:
            wb.save('releasesData.xlsx')
            return i
        else:
            wb.save('releasesData.xlsx')
            i += 1
    return i


def insert_release_to_db(release):
    """
    This function enter to excel file new release save in both sheets 'Releases' and 'Release notes'
    :param release:
    :return:
    """
    wb = openpyxl.load_workbook('releasesData.xlsx')
    sheet = wb['Releases']

    id = release['ID']
    release_notes = release['Release notes']
    release_date = release['Release Date']
    author = release['Author']

    last_raw = find_last_row(id)
    if last_raw < 0:
        return False
    sheet.cell(row=last_raw, column=1).value = id
    sheet.cell(row=last_raw, column=2).value = author
    sheet.cell(row=last_raw, column=3).value = release_date
    wb.save('releasesData.xlsx')

    sheet = wb['Release notes']
    i = sheet.max_row+1
    for note in release_notes:
        sheet.cell(row=i, column=1).value = id
        sheet.cell(row=i, column=2).value = note
        i += 1

    wb.save('releasesData.xlsx')
    return True


def init_versions():
    """
    This function return a list of existing versions found in excel file
    :return:
    """

    release_file = pd.read_excel('releasesData.xlsx', sheet_name='Releases')
    VERSION = []
    for index, version in release_file.iterrows():
        id = version[0]
        author = version[1]
        release_date = version[2]
        notes = get_notes_of_release(release_id=id)

        VERSION.append({
            'ID': id,
            'Release notes': notes,
            'Release Date': release_date,
            'Author': author
        })
    return VERSION


def get_notes_of_release(release_id):
    """
    This function returns an array of notes
    :param release_id:
    :return:
    """
    release_notes_file = pd.read_excel('releasesData.xlsx', sheet_name='Release notes')
    releases_for_id = release_notes_file.loc[release_notes_file['ReleaseID'] == release_id]
    notes = [note[1] for index, note in releases_for_id.iterrows()]
    return notes


def get_release_request(post_data):
    """
    This function return an object release from a request data
    :param post_data:
    :return:
    """

    release = {
        'ID': post_data.get('id'),
        'Release notes': post_data.get('releaseNotes'),
        'Release Date': post_data.get('releaseDate'),
        'Author': post_data.get('author')
    }
    return release


def delete_release(release_id):
    """
    This function delete a release from excel file
    and saving changes in the file.
    :param release_id:
    :return:
    """
    wb = openpyxl.load_workbook('releasesData.xlsx')
    delete_from_sheet(wb['Release notes'], release_id)
    wb.save('releasesData.xlsx')
    delete_from_sheet(wb['Releases'], release_id)
    wb.save('releasesData.xlsx')


def delete_from_sheet(sheet, release_id):
    """
    This function delete release by id from sheet in excel file
    :param sheet:
    :param release_id:
    :return:
    """
    i = 1
    while i <= sheet.max_row:
        if sheet.cell(row=i, column=1).value == release_id:
            sheet.delete_rows(i, 1)
        else:
            i += 1


if __name__ == '__main__':
    app.run()
